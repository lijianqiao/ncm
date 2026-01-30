"""
@Author: li
@Email: lijianqiao2906@live.com
@FileName: parse.py
@DateTime: 2026/01/16 08:40:18
@Docs: 解析表格文件，支持CSV和Excel格式
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

import polars as pl
from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class ParsedTable:
    """解析后的表格数据。

    Attributes:
        df (pl.DataFrame): 解析后的 DataFrame。
        total_rows (int): 总行数。
        columns (list[str]): 列名列表。
    """

    df: pl.DataFrame
    total_rows: int
    columns: list[str]


def _read_excel_to_polars(file_path: Path, *, sheet_name: str | int | None = None) -> pl.DataFrame:
    """读取 Excel 文件为 Polars DataFrame。

    Args:
        file_path (Path): Excel 文件路径。
        sheet_name (str | int | None): 工作表名称或索引，默认 None（第一个工作表）。

    Returns:
        pl.DataFrame: 解析后的 DataFrame。
    """
    try:
        read_excel = getattr(pl, "read_excel", None)
        if callable(read_excel):
            df = read_excel(str(file_path), sheet_name=sheet_name)  # type: ignore[call-arg]
            return cast(pl.DataFrame, df)
    except Exception:
        pass
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            ws = wb.worksheets[0]
        elif isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return pl.DataFrame()

        headers: list[str] = []
        seen: dict[str, int] = {}
        for idx, h in enumerate(header_row, start=1):
            name = str(h).strip() if h is not None and str(h).strip() else f"column_{idx}"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            headers.append(name)

        records: list[dict[str, Any]] = []
        for row in rows_iter:
            if row is None:
                continue
            record = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            records.append(record)

        return pl.DataFrame(records)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def parse_tabular_file(file_path: Path, *, filename: str) -> ParsedTable:
    """解析表格文件（CSV 或 Excel）。

    Args:
        file_path (Path): 文件路径。
        filename (str): 文件名（用于确定文件类型）。

    Returns:
        ParsedTable: 解析后的表格数据。

    Raises:
        ValueError: 当文件类型不支持时。
    """
    suffix = Path(filename).suffix.lower().lstrip(".")
    if suffix in {"csv"}:
        df = pl.read_csv(str(file_path), infer_schema=False, encoding="utf8-lossy")
    elif suffix in {"xlsx", "xlsm", "xls"}:
        df = _read_excel_to_polars(file_path)
    else:
        raise ValueError(f"不支持的文件类型: .{suffix}")

    df = df.with_columns([pl.all().cast(pl.Utf8, strict=False)])
    df = df.with_row_index(name="row_number", offset=1)

    return ParsedTable(df=df, total_rows=df.height, columns=list(df.columns))


def normalize_columns(df: pl.DataFrame, column_mapping: dict[str, str]) -> pl.DataFrame:
    """规范化 DataFrame 列名。

    Args:
        df (pl.DataFrame): 原始 DataFrame。
        column_mapping (dict[str, str]): 列名映射字典（旧列名 -> 新列名）。

    Returns:
        pl.DataFrame: 列名规范化后的 DataFrame。
    """
    normalized: dict[str, str] = {}
    for c in df.columns:
        c_norm = str(c).strip()
        normalized[c] = column_mapping.get(c_norm, c_norm)
    return df.rename(normalized)


def dataframe_to_preview_rows(df: pl.DataFrame) -> list[dict[str, Any]]:
    """将 DataFrame 转换为预览行列表。

    Args:
        df (pl.DataFrame): DataFrame。

    Returns:
        list[dict[str, Any]]: 行字典列表。
    """
    return df.to_dicts()
