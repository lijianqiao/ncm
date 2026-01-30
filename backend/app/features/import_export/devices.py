"""
@Author: li
@Email: lijianqiao2906@live.com
@FileName: devices.py
@DateTime: 2026/01/16 08:41:09
@Docs: 设备导入导出
"""

import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, cast
from uuid import UUID

import polars as pl
import uuid6
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import and_, insert, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.encryption import encrypt_credential
from app.core.enums import AuthType, DeviceGroup, DeviceStatus, DeviceVendor
from app.models.credential import DeviceGroupCredential
from app.models.dept import Department
from app.models.device import Device
from fastapi_import_export.validation_extras import RowValidator

DEVICE_IMPORT_COLUMNS: list[str] = [
    "name",
    "ip_address",
    "vendor",
    "model",
    "platform",
    "location",
    "description",
    "ssh_port",
    "auth_type",
    "dept_code",
    "device_group",
    "status",
    "username",
    "password",
    "serial_number",
    "os_version",
    "stock_in_at",
    "assigned_to",
]

DEVICE_IMPORT_OPTIONAL_COLUMNS: list[str] = [
    "otp_seed",
]

DEVICE_IMPORT_TEMPLATE_COLUMNS: list[str] = [
    *DEVICE_IMPORT_COLUMNS,
    *DEVICE_IMPORT_OPTIONAL_COLUMNS,
]

DEVICE_IMPORT_COLUMN_ALIASES: dict[str, str] = {
    "设备名称": "name",
    "名称": "name",
    "IP": "ip_address",
    "IP地址": "ip_address",
    "ip": "ip_address",
    "厂商": "vendor",
    "型号": "model",
    "平台": "platform",
    "位置": "location",
    "描述": "description",
    "SSH端口": "ssh_port",
    "认证类型": "auth_type",
    "部门编码": "dept_code",
    "设备分组": "device_group",
    "状态": "status",
    "用户名": "username",
    "密码": "password",
    "序列号": "serial_number",
    "系统版本": "os_version",
    "入库时间": "stock_in_at",
    "领用人": "assigned_to",
    "OTP种子": "otp_seed",
    "OTP 种子": "otp_seed",
    "otp_seed": "otp_seed",
}


def _to_str(v: Any) -> str:
    """将值转换为字符串。

    Args:
        v (Any): 输入值。

    Returns:
        str: 转换后的字符串，如果输入为 None 则返回空字符串。
    """
    if v is None:
        return ""
    return str(v).strip()


def _normalize_enum(value: str) -> str:
    """规范化枚举值。

    Args:
        value (str): 枚举值字符串。

    Returns:
        str: 规范化后的枚举值（小写，空格替换为下划线）。
    """
    return value.strip().lower().replace(" ", "_")


def _parse_int(value: str, default: int | None = None) -> int | None:
    """解析整数。

    Args:
        value (str): 字符串值。
        default (int | None): 默认值，当解析失败或为空时返回。

    Returns:
        int | None: 解析后的整数值，失败时返回默认值。
    """
    v = value.strip()
    if not v:
        return default
    try:
        return int(float(v))
    except Exception:
        return None


def _parse_datetime(value: str) -> datetime | None:
    """解析日期时间字符串。

    Args:
        value (str): 日期时间字符串。

    Returns:
        datetime | None: 解析后的日期时间对象，失败时返回 None。
    """
    v = value.strip()
    if not v:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(v, fmt)
        except Exception:
            continue
    try:
        return datetime.fromisoformat(v)
    except Exception:
        return None


async def validate_devices(
    db: AsyncSession,
    df: pl.DataFrame,
    *,
    allow_overwrite: bool = False,
) -> tuple[pl.DataFrame, list[dict[str, Any]]]:
    """验证设备导入数据。

    Args:
        db (AsyncSession): 数据库会话。
        df (pl.DataFrame): 设备数据 DataFrame。
        allow_overwrite (bool): 是否允许覆盖已存在的设备，默认 False。

    Returns:
        tuple[pl.DataFrame, list[dict[str, Any]]]: 验证后的有效数据 DataFrame 和错误列表。
    """
    errors: list[dict[str, Any]] = []

    for col in DEVICE_IMPORT_COLUMNS:
        if col not in df.columns:
            errors.append({"row_number": 0, "field": col, "message": f"缺少列: {col}"})

    if errors:
        return pl.DataFrame(), errors

    def add_error(row_number: int, field: str | None, message: str) -> None:
        errors.append({"row_number": int(row_number), "field": field, "message": message})

    select_cols = [*DEVICE_IMPORT_COLUMNS, "row_number"]
    if "otp_seed" in df.columns:
        select_cols.append("otp_seed")
    rows = df.select(select_cols).to_dicts()

    ips = [_to_str(r.get("ip_address")) for r in rows]
    ip_set = {ip for ip in ips if ip}

    result = await db.execute(select(Device.ip_address, Device.is_deleted).where(Device.ip_address.in_(ip_set)))
    existing_ip_deleted_map = {ip: bool(is_deleted) for ip, is_deleted in result.all()}

    dept_codes = {_to_str(r.get("dept_code")) for r in rows if _to_str(r.get("dept_code"))}
    dept_map: dict[str, UUID] = {}
    if dept_codes:
        dept_result = await db.execute(select(Department.code, Department.id).where(Department.code.in_(dept_codes)))
        dept_map = {str(code): dept_id for code, dept_id in dept_result.all()}

    credential_pairs: set[tuple[UUID, str]] = set()
    credential_seed_pairs: set[tuple[UUID, str]] = set()

    ip_seen: set[str] = set()
    valid_row_numbers: list[int] = []
    normalized_rows: list[dict[str, Any]] = []

    for r in rows:
        row_number = int(r["row_number"])
        vendor = _normalize_enum(_to_str(r.get("vendor")) or DeviceVendor.H3C.value)
        auth_type = _normalize_enum(_to_str(r.get("auth_type")) or AuthType.OTP_SEED.value)
        device_group = _normalize_enum(_to_str(r.get("device_group")) or DeviceGroup.ACCESS.value)
        status = _normalize_enum(_to_str(r.get("status")) or DeviceStatus.IN_USE.value)
        dept_code = _to_str(r.get("dept_code"))

        r["vendor"] = vendor
        r["auth_type"] = auth_type
        r["device_group"] = device_group
        r["status"] = status

        v = RowValidator(errors=errors, row_number=row_number, row=r)
        v.not_blank("name", "设备名称不能为空")
        v.not_blank("ip_address", "IP 地址不能为空")

        ip = v.get_str("ip_address")
        if ip:
            if ip in ip_seen:
                v.add(field="ip_address", message=f"导入文件中 IP 地址重复: {ip}", value=ip, type="infile_duplicate")
            ip_seen.add(ip)
            v.ip_address("ip_address", "IP 地址格式无效")
            v.db_unique_conflict(
                field="ip_address",
                deleted_map=existing_ip_deleted_map,
                allow_overwrite=allow_overwrite,
                exists_message="IP 地址已存在: {value}",
                deleted_message="IP 地址已存在（回收站）: {value}，请先恢复或选择覆盖导入",
            )

        v.one_of("vendor", {e.value for e in DeviceVendor}, "厂商无效")
        v.one_of("auth_type", {e.value for e in AuthType}, "认证类型无效")
        v.one_of("device_group", {e.value for e in DeviceGroup}, "设备分组无效")
        v.one_of("status", {e.value for e in DeviceStatus}, "设备状态无效")

        dept_id: UUID | None = None
        if dept_code:
            dept_id = dept_map.get(dept_code)
            if dept_id is None:
                add_error(row_number, "dept_code", f"部门不存在: {dept_code}")

        if auth_type in {AuthType.OTP_SEED.value, AuthType.OTP_MANUAL.value}:
            if not dept_code:
                add_error(row_number, "dept_code", "OTP 认证设备必须填写部门编码 dept_code")
            elif dept_id is not None:
                credential_pairs.add((dept_id, device_group))
                otp_seed_value = _to_str(r.get("otp_seed"))
                otp_username_value = _to_str(r.get("username"))
                if auth_type == AuthType.OTP_SEED.value and otp_seed_value and otp_username_value:
                    credential_seed_pairs.add((dept_id, device_group))

        if auth_type == AuthType.STATIC.value:
            v.require_fields(["username", "password"], "静态密码认证必须填写")

        normalized_rows.append(
            {
                "row_number": row_number,
                "name": _to_str(r.get("name")),
                "ip_address": ip,
                "vendor": vendor,
                "model": _to_str(r.get("model")) or None,
                "platform": _to_str(r.get("platform")) or None,
                "location": _to_str(r.get("location")) or None,
                "description": _to_str(r.get("description")) or None,
                "ssh_port": _parse_int(_to_str(r.get("ssh_port")), default=22),
                "auth_type": auth_type,
                "dept_code": dept_code or None,
                "dept_id": str(dept_id) if dept_id else None,
                "device_group": device_group,
                "status": status,
                "username": _to_str(r.get("username")) or None,
                "password": _to_str(r.get("password")) or None,
                "otp_seed": _to_str(r.get("otp_seed")) or None,
                "serial_number": _to_str(r.get("serial_number")) or None,
                "os_version": _to_str(r.get("os_version")) or None,
                "stock_in_at": (_parse_datetime(_to_str(r.get("stock_in_at"))) or None),
                "assigned_to": _to_str(r.get("assigned_to")) or None,
            }
        )

    if credential_pairs:
        dept_ids = {dept_id for dept_id, _ in credential_pairs}
        groups = {g for _, g in credential_pairs}
        cred_result = await db.execute(
            select(DeviceGroupCredential.dept_id, DeviceGroupCredential.device_group).where(
                and_(DeviceGroupCredential.dept_id.in_(dept_ids), DeviceGroupCredential.device_group.in_(groups))
            )
        )
        existing_pairs = {(dept_id, group) for dept_id, group in cred_result.all()}
        for dept_id, group in credential_pairs:
            if (dept_id, group) not in existing_pairs:
                if (dept_id, group) in credential_seed_pairs:
                    continue
                missing_codes = [code for code, did in dept_map.items() if did == dept_id]
                code_display = missing_codes[0] if missing_codes else str(dept_id)
                errors.append(
                    {
                        "row_number": 0,
                        "field": "dept_code",
                        "message": f"部门 {code_display} 分组 {group} 缺少凭据（请先创建凭据）",
                    }
                )

    error_row_numbers = {int(e["row_number"]) for e in errors if int(e["row_number"]) > 0}
    for r in normalized_rows:
        if r["row_number"] not in error_row_numbers:
            valid_row_numbers.append(r["row_number"])

    valid_rows = [r for r in normalized_rows if r["row_number"] in set(valid_row_numbers)]
    for r in valid_rows:
        if isinstance(r.get("stock_in_at"), datetime):
            r["stock_in_at"] = r["stock_in_at"].isoformat()
    valid_df = pl.DataFrame(valid_rows)
    return valid_df, errors


async def persist_devices(
    db: AsyncSession,
    valid_df: pl.DataFrame,
    *,
    allow_overwrite: bool = False,
) -> int:
    """持久化设备数据。

    Args:
        db (AsyncSession): 数据库会话。
        valid_df (pl.DataFrame): 已验证的设备数据 DataFrame。
        allow_overwrite (bool): 是否允许覆盖已存在的设备，默认 False。

    Returns:
        int: 成功持久化的设备数量。
    """
    if valid_df.is_empty():
        return 0

    device_dicts = valid_df.to_dicts()
    now = datetime.now().astimezone().replace(tzinfo=None)

    async with db.begin():
        ip_set = {str(d["ip_address"]) for d in device_dicts}
        existing_result = await db.execute(select(Device).where(Device.ip_address.in_(ip_set)))
        existing_by_ip = {d.ip_address: d for d in existing_result.scalars().all()}

        to_insert: list[dict[str, Any]] = []
        to_update: list[tuple[Device, dict[str, Any]]] = []

        credential_seed_items: list[tuple[UUID, str, str, str]] = []

        for d in device_dicts:
            ip = str(d["ip_address"])
            auth_type = str(d["auth_type"])
            dept_id: UUID | None = None
            if d.get("dept_id"):
                try:
                    dept_id = UUID(str(d["dept_id"]))
                except Exception:
                    dept_id = None
            stock_in_at: datetime | None = None
            if d.get("stock_in_at"):
                stock_in_at = _parse_datetime(str(d["stock_in_at"])) or None
            password_encrypted: str | None = None
            if auth_type == AuthType.STATIC.value and d.get("password"):
                password_encrypted = encrypt_credential(str(d["password"]), settings.NCM_CREDENTIAL_KEY)

            otp_seed: str | None = None
            if d.get("otp_seed"):
                otp_seed = str(d["otp_seed"]).strip() or None
            otp_username: str | None = None
            if d.get("username"):
                otp_username = str(d["username"]).strip() or None

            if (
                auth_type == AuthType.OTP_SEED.value
                and dept_id is not None
                and otp_seed
                and otp_username
                and d.get("device_group")
            ):
                credential_seed_items.append((dept_id, str(d.get("device_group")), otp_username, otp_seed))

            values: dict[str, Any] = {
                "name": d.get("name"),
                "ip_address": ip,
                "vendor": d.get("vendor"),
                "model": d.get("model"),
                "platform": d.get("platform"),
                "location": d.get("location"),
                "description": d.get("description"),
                "ssh_port": int(d.get("ssh_port") or 22),
                "auth_type": auth_type,
                "dept_id": dept_id,
                "device_group": d.get("device_group"),
                "status": d.get("status"),
                "username": d.get("username") if auth_type == AuthType.STATIC.value else None,
                "password_encrypted": password_encrypted,
                "serial_number": d.get("serial_number"),
                "os_version": d.get("os_version"),
                "stock_in_at": stock_in_at,
                "assigned_to": d.get("assigned_to"),
                "updated_at": now,
            }

            existing = existing_by_ip.get(ip)
            if existing is not None:
                if allow_overwrite:
                    if existing.is_deleted:
                        existing.is_deleted = False
                        existing.is_active = True
                    to_update.append((existing, values))
                else:
                    continue
            else:
                values["id"] = uuid6.uuid7()
                values["version_id"] = uuid.uuid4().hex
                values["created_at"] = now
                to_insert.append(values)

        if to_update:
            for existing, values in to_update:
                for k, v in values.items():
                    setattr(existing, k, v)
        if to_insert:
            await db.execute(insert(Device), to_insert)

        if credential_seed_items:
            pairs = {(dept_id, group) for dept_id, group, _, _ in credential_seed_items}
            dept_ids = {dept_id for dept_id, _ in pairs}
            groups = {group for _, group in pairs}
            existing_cred_result = await db.execute(
                select(DeviceGroupCredential.dept_id, DeviceGroupCredential.device_group).where(
                    and_(DeviceGroupCredential.dept_id.in_(dept_ids), DeviceGroupCredential.device_group.in_(groups))
                )
            )
            existing_pairs = {(dept_id, group) for dept_id, group in existing_cred_result.all()}
            to_create: list[dict[str, Any]] = []
            for dept_id, group, username, seed in credential_seed_items:
                if (dept_id, group) in existing_pairs:
                    continue
                to_create.append(
                    {
                        "id": uuid6.uuid7(),
                        "version_id": uuid.uuid4().hex,
                        "created_at": now,
                        "updated_at": now,
                        "dept_id": dept_id,
                        "device_group": group,
                        "username": username,
                        "otp_seed_encrypted": encrypt_credential(seed, settings.NCM_CREDENTIAL_KEY),
                        "auth_type": AuthType.OTP_SEED.value,
                        "description": "imported",
                    }
                )
            if to_create:
                await db.execute(insert(DeviceGroupCredential), to_create)

    return len(to_insert) + len(to_update)


async def export_devices_df(db: AsyncSession) -> pl.DataFrame:
    """导出设备数据为 DataFrame。

    Args:
        db (AsyncSession): 数据库会话。

    Returns:
        pl.DataFrame: 设备数据 DataFrame。
    """
    result = await db.execute(select(Device).where(Device.is_deleted.is_(False)).order_by(Device.created_at.desc()))
    devices = result.scalars().all()
    rows: list[dict[str, Any]] = []
    for d in devices:
        rows.append(
            {
                "name": d.name,
                "ip_address": d.ip_address,
                "vendor": d.vendor,
                "model": d.model,
                "platform": d.platform,
                "location": d.location,
                "ssh_port": d.ssh_port,
                "auth_type": d.auth_type,
                "dept_id": str(d.dept_id) if d.dept_id else "",
                "device_group": d.device_group,
                "status": d.status,
                "serial_number": d.serial_number or "",
                "os_version": d.os_version or "",
                "stock_in_at": d.stock_in_at.isoformat() if d.stock_in_at else "",
                "assigned_to": d.assigned_to or "",
                "created_at": d.created_at.isoformat() if d.created_at else "",
                "updated_at": d.updated_at.isoformat() if d.updated_at else "",
            }
        )
    return pl.DataFrame(rows)


def build_device_import_template(path: Path) -> None:
    """构建设备导入模板 Excel 文件。

    Args:
        path (Path): 模板文件保存路径。

    Raises:
        RuntimeError: 当 Workbook.active 为 None 时。
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook.active is None")
    ws = cast(Worksheet, ws)
    ws.title = "devices"

    header_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(DEVICE_IMPORT_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(12, min(24, len(col_name) + 4))

    samples = [
        {
            "name": "core-switch-01",
            "ip_address": "192.168.10.1",
            "vendor": "h3c",
            "model": "S5500",
            "platform": "hp_comware",
            "location": "机房A",
            "description": "示例设备（OTP 种子）",
            "ssh_port": "22",
            "auth_type": "otp_seed",
            "dept_code": "HQ",
            "device_group": "core",
            "status": "in_use",
            "username": "admin",
            "password": "",
            "otp_seed": "JBSWY3DPEHPK3PXP",
            "serial_number": "",
            "os_version": "",
            "stock_in_at": "2026-01-01",
            "assigned_to": "",
        },
        {
            "name": "access-switch-01",
            "ip_address": "192.168.20.10",
            "vendor": "huawei",
            "model": "S5735",
            "platform": "huawei_vrp",
            "location": "机房B",
            "description": "示例设备（静态密码）",
            "ssh_port": "22",
            "auth_type": "static",
            "dept_code": "",
            "device_group": "access",
            "status": "active",
            "username": "admin",
            "password": "Password123!",
            "otp_seed": "",
            "serial_number": "SN-001",
            "os_version": "V200R020",
            "stock_in_at": "",
            "assigned_to": "",
        },
        {
            "name": "distribution-switch-01",
            "ip_address": "192.168.30.2",
            "vendor": "cisco",
            "model": "C9300",
            "platform": "cisco_iosxe",
            "location": "机房C",
            "description": "示例设备（OTP 手动输入）",
            "ssh_port": "22",
            "auth_type": "otp_manual",
            "dept_code": "HQ",
            "device_group": "distribution",
            "status": "maintenance",
            "username": "admin",
            "password": "",
            "otp_seed": "",
            "serial_number": "SN-9300",
            "os_version": "17.09",
            "stock_in_at": "",
            "assigned_to": "张三",
        },
    ]

    for row_idx, sample in enumerate(samples, start=2):
        for col_idx, col_name in enumerate(DEVICE_IMPORT_TEMPLATE_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=sample.get(col_name, ""))

    dv_vendor = DataValidation(type="list", formula1=f'"{",".join([e.value for e in DeviceVendor])}"', allow_blank=True)
    dv_group = DataValidation(type="list", formula1=f'"{",".join([e.value for e in DeviceGroup])}"', allow_blank=True)
    dv_auth = DataValidation(type="list", formula1=f'"{",".join([e.value for e in AuthType])}"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1=f'"{",".join([e.value for e in DeviceStatus])}"', allow_blank=True)

    vendor_col = DEVICE_IMPORT_COLUMNS.index("vendor") + 1
    group_col = DEVICE_IMPORT_COLUMNS.index("device_group") + 1
    auth_col = DEVICE_IMPORT_COLUMNS.index("auth_type") + 1
    status_col = DEVICE_IMPORT_COLUMNS.index("status") + 1

    for dv, col in ((dv_vendor, vendor_col), (dv_group, group_col), (dv_auth, auth_col), (dv_status, status_col)):
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=col).coordinate}:{ws.cell(row=500, column=col).coordinate}")

    ws.freeze_panes = "A2"
    wb.save(path)
