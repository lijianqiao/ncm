"""
@Author: li
@Email: li
@FileName: inventory_audit.py
@DateTime: 2026/01/16
@Docs: 资产盘点导出
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.discovery import Discovery
from app.models.device import Device
from app.models.inventory_audit import InventoryAudit
from app.utils.user_display import format_user_display_name

# 状态中英文映射
STATUS_MAP = {
    "pending": "待处理",
    "matched": "已匹配",
    "shadow": "影子资产",
    "adopted": "已纳管",
    "ignored": "已忽略",
    "success": "成功",
    "running": "运行中",
    "failed": "失败",
    "partial": "部分成功",
}

# 样式定义
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12)
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _format_datetime(dt: datetime | None) -> str:
    """格式化日期时间。"""
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _format_ports(ports: dict | None) -> str:
    """格式化端口信息。"""
    if not ports:
        return ""
    return ", ".join(f"{p}({s})" for p, s in ports.items())


def _apply_header_style(ws: Worksheet, row: int, col_count: int) -> None:
    """应用表头样式。"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_column_width(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
    """自动调整列宽。"""
    for col_cells in ws.columns:
        max_len = 0
        col_idx = col_cells[0].column
        if col_idx is None:
            continue
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                # 中文字符按 2 个宽度计算
                for char in str(cell.value or ""):
                    if "\u4e00" <= char <= "\u9fff":
                        cell_len += 1
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        adjusted_width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


async def export_inventory_audits_df(db: AsyncSession) -> pl.DataFrame:
    """导出盘点任务列表（用于任务列表导出）。"""
    result = await db.execute(
        select(InventoryAudit).where(InventoryAudit.is_deleted.is_(False)).order_by(InventoryAudit.created_at.desc())
    )
    rows: list[dict[str, Any]] = []
    for a in result.scalars().all():
        operator_name = ""
        if a.operator:
            operator_name = format_user_display_name(a.operator.nickname, a.operator.username) or ""
        rows.append(
            {
                "任务名称": a.name,
                "状态": STATUS_MAP.get(a.status, a.status),
                "扫描范围": json.dumps(a.scope or {}, ensure_ascii=False),
                "操作人": operator_name,
                "开始时间": _format_datetime(a.started_at),
                "结束时间": _format_datetime(a.finished_at),
                "创建时间": _format_datetime(a.created_at),
            }
        )
    return pl.DataFrame(rows)


async def export_inventory_audit_report(
    db: AsyncSession,
    audit_id: str,
    output_path: Path,
) -> Path:
    """
    生成单个盘点任务的详细 Excel 报告。

    包含多个 Sheet：
    - 盘点概要：任务信息、统计数据
    - 发现设备明细：所有扫描到的设备
    - 影子资产：未在 CMDB 中的设备
    - 离线设备：CMDB 有但扫描未发现的设备

    Args:
        db: 数据库会话
        audit_id: 盘点任务 ID
        output_path: 输出文件路径

    Returns:
        输出文件路径
    """
    from uuid import UUID

    # 1. 获取盘点任务
    audit = await db.get(InventoryAudit, UUID(audit_id))
    if not audit:
        raise ValueError("盘点任务不存在")

    # 2. 获取本次盘点发现的所有设备
    discoveries_result = await db.execute(
        select(Discovery)
        .where(Discovery.scan_task_id == audit_id)
        .order_by(Discovery.ip_address)
    )
    discoveries = list(discoveries_result.scalars().all())

    # 3. 获取离线设备（CMDB 有但扫描未发现）
    # 获取本次扫描发现的所有 IP
    discovered_ips = {d.ip_address for d in discoveries}

    # 解析扫描范围
    scope = audit.scope or {}
    subnets = scope.get("subnets") or []

    # 查询 CMDB 中应该被扫描到的设备
    offline_devices: list[Device] = []
    if subnets:
        # 简化处理：获取所有活跃设备，检查是否在发现列表中
        # 显式加载 dept 关系，避免异步上下文中的延迟加载问题
        devices_result = await db.execute(
            select(Device).where(Device.is_deleted.is_(False)).options(selectinload(Device.dept))
        )
        all_devices = list(devices_result.scalars().all())
        for device in all_devices:
            if device.ip_address and device.ip_address not in discovered_ips:
                # 检查设备 IP 是否在扫描范围内
                import ipaddress

                for subnet in subnets:
                    try:
                        network = ipaddress.ip_network(subnet, strict=False)
                        if ipaddress.ip_address(device.ip_address) in network:
                            offline_devices.append(device)
                            break
                    except Exception:
                        pass

    # 4. 创建 Excel 工作簿
    wb = Workbook()

    # ========== Sheet1: 盘点概要 ==========
    ws_summary: Worksheet = wb.active  # type: ignore[assignment]
    ws_summary.title = "盘点概要"

    # 标题
    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = "资产盘点报告"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGNMENT

    # 任务信息
    info_title = ws_summary["A3"]
    info_title.value = "任务信息"
    info_title.font = SUBTITLE_FONT

    task_info = [
        ("任务名称", audit.name),
        ("任务状态", STATUS_MAP.get(audit.status, audit.status)),
        ("扫描范围", ", ".join(subnets) if subnets else "未指定"),
        ("操作人", format_user_display_name(audit.operator.nickname, audit.operator.username) if audit.operator else ""),
        ("开始时间", _format_datetime(audit.started_at)),
        ("结束时间", _format_datetime(audit.finished_at)),
    ]

    for i, (label, value) in enumerate(task_info, start=4):
        label_cell = ws_summary.cell(row=i, column=1, value=label)
        label_cell.font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=value)

    # 统计数据
    result = audit.result or {}
    cmdb_compare = result.get("cmdb_compare") or {}

    stats_title = ws_summary.cell(row=11, column=1, value="统计数据")
    stats_title.font = SUBTITLE_FONT

    stats = [
        ("发现设备总数", result.get("discoveries_total", len(discoveries))),
        ("CMDB 设备总数", cmdb_compare.get("total_cmdb", 0)),
        ("已匹配设备", cmdb_compare.get("matched", 0)),
        ("影子资产（未纳管）", cmdb_compare.get("shadow_assets", 0)),
        ("离线设备（未扫描到）", cmdb_compare.get("offline_devices", len(offline_devices))),
    ]

    for i, (label, value) in enumerate(stats, start=12):
        label_cell = ws_summary.cell(row=i, column=1, value=label)
        label_cell.font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=value)

    # 状态分布
    by_status = result.get("discoveries_by_status") or {}
    status_title = ws_summary.cell(row=18, column=1, value="发现设备状态分布")
    status_title.font = SUBTITLE_FONT

    row = 19
    for status, count in by_status.items():
        ws_summary.cell(row=row, column=1, value=STATUS_MAP.get(status, status))
        ws_summary.cell(row=row, column=2, value=count)
        row += 1

    _auto_column_width(ws_summary)

    # ========== Sheet2: 发现设备明细 ==========
    ws_discoveries = wb.create_sheet("发现设备明细")

    discovery_headers = [
        "IP地址", "MAC地址", "主机名", "厂商", "操作系统",
        "开放端口", "发现状态", "首次发现", "最后发现", "离线天数"
    ]
    for col, header in enumerate(discovery_headers, start=1):
        ws_discoveries.cell(row=1, column=col, value=header)
    _apply_header_style(ws_discoveries, 1, len(discovery_headers))

    for row_idx, d in enumerate(discoveries, start=2):
        ws_discoveries.cell(row=row_idx, column=1, value=d.ip_address)
        ws_discoveries.cell(row=row_idx, column=2, value=d.mac_address or "")
        ws_discoveries.cell(row=row_idx, column=3, value=d.hostname or d.snmp_sysname or "")
        ws_discoveries.cell(row=row_idx, column=4, value=d.vendor or "")
        ws_discoveries.cell(row=row_idx, column=5, value=d.os_info or "")
        ws_discoveries.cell(row=row_idx, column=6, value=_format_ports(d.open_ports))
        ws_discoveries.cell(row=row_idx, column=7, value=STATUS_MAP.get(d.status, d.status))
        ws_discoveries.cell(row=row_idx, column=8, value=_format_datetime(d.first_seen_at))
        ws_discoveries.cell(row=row_idx, column=9, value=_format_datetime(d.last_seen_at))
        ws_discoveries.cell(row=row_idx, column=10, value=d.offline_days)

    _auto_column_width(ws_discoveries)

    # ========== Sheet3: 影子资产 ==========
    ws_shadow = wb.create_sheet("影子资产")

    shadow_devices = [d for d in discoveries if d.status == "shadow"]

    shadow_headers = [
        "IP地址", "MAC地址", "主机名", "厂商", "操作系统",
        "开放端口", "首次发现", "建议操作"
    ]
    for col, header in enumerate(shadow_headers, start=1):
        ws_shadow.cell(row=1, column=col, value=header)
    _apply_header_style(ws_shadow, 1, len(shadow_headers))

    for row_idx, d in enumerate(shadow_devices, start=2):
        ws_shadow.cell(row=row_idx, column=1, value=d.ip_address)
        ws_shadow.cell(row=row_idx, column=2, value=d.mac_address or "")
        ws_shadow.cell(row=row_idx, column=3, value=d.hostname or d.snmp_sysname or "")
        ws_shadow.cell(row=row_idx, column=4, value=d.vendor or "")
        ws_shadow.cell(row=row_idx, column=5, value=d.os_info or "")
        ws_shadow.cell(row=row_idx, column=6, value=_format_ports(d.open_ports))
        ws_shadow.cell(row=row_idx, column=7, value=_format_datetime(d.first_seen_at))
        ws_shadow.cell(row=row_idx, column=8, value="建议纳管或标记忽略")

    if not shadow_devices:
        ws_shadow.cell(row=2, column=1, value="无影子资产")

    _auto_column_width(ws_shadow)

    # ========== Sheet4: 离线设备 ==========
    ws_offline = wb.create_sheet("离线设备")

    offline_headers = [
        "设备名称", "IP地址", "厂商", "设备分组", "部门",
        "状态", "建议操作"
    ]
    for col, header in enumerate(offline_headers, start=1):
        ws_offline.cell(row=1, column=col, value=header)
    _apply_header_style(ws_offline, 1, len(offline_headers))

    for row_idx, device in enumerate(offline_devices, start=2):
        ws_offline.cell(row=row_idx, column=1, value=device.name)
        ws_offline.cell(row=row_idx, column=2, value=device.ip_address)
        ws_offline.cell(row=row_idx, column=3, value=device.vendor or "")
        ws_offline.cell(row=row_idx, column=4, value=device.device_group or "")
        ws_offline.cell(row=row_idx, column=5, value=device.dept.name if device.dept else "")
        ws_offline.cell(row=row_idx, column=6, value=device.status or "")
        ws_offline.cell(row=row_idx, column=7, value="检查设备是否在线或已下架")

    if not offline_devices:
        ws_offline.cell(row=2, column=1, value="无离线设备")

    _auto_column_width(ws_offline)

    # 5. 保存文件（使用字符串路径确保兼容性）
    try:
        wb.save(str(output_path))
    finally:
        wb.close()
    return output_path

