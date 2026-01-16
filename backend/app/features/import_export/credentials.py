"""
@Author: li
@Email: lijianqiao2906@live.com
@FileName: credentials.py
@DateTime: 2026/01/16
@Docs: 分组凭据导入导出（部门 + 设备分组）
"""

import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, cast
from uuid import UUID

import polars as pl
import uuid6
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import and_, insert, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.encryption import encrypt_credential
from app.core.enums import AuthType, DeviceGroup
from app.models.credential import DeviceGroupCredential
from app.models.dept import Department

CREDENTIAL_IMPORT_COLUMNS: list[str] = [
    "dept_code",
    "device_group",
    "auth_type",
    "username",
    "description",
]

CREDENTIAL_IMPORT_OPTIONAL_COLUMNS: list[str] = [
    "otp_seed",
]

CREDENTIAL_IMPORT_TEMPLATE_COLUMNS: list[str] = [
    *CREDENTIAL_IMPORT_COLUMNS,
    *CREDENTIAL_IMPORT_OPTIONAL_COLUMNS,
]

CREDENTIAL_IMPORT_COLUMN_ALIASES: dict[str, str] = {
    "部门编码": "dept_code",
    "部门": "dept_code",
    "dept_code": "dept_code",
    "分组": "device_group",
    "设备分组": "device_group",
    "device_group": "device_group",
    "认证类型": "auth_type",
    "auth_type": "auth_type",
    "用户名": "username",
    "username": "username",
    "描述": "description",
    "说明": "description",
    "description": "description",
    "OTP种子": "otp_seed",
    "OTP 种子": "otp_seed",
    "otp_seed": "otp_seed",
}


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize_enum(value: str) -> str:
    return value.strip().lower().replace(" ", "_")


async def validate_credentials(
    db: AsyncSession,
    df: pl.DataFrame,
    *,
    allow_overwrite: bool = False,
) -> tuple[pl.DataFrame, list[dict[str, Any]]]:
    errors: list[dict[str, Any]] = []

    for col in CREDENTIAL_IMPORT_COLUMNS:
        if col not in df.columns:
            errors.append({"row_number": 0, "field": col, "message": f"缺少列: {col}"})

    if errors:
        return pl.DataFrame(), errors

    def add_error(row_number: int, field: str | None, message: str) -> None:
        errors.append({"row_number": int(row_number), "field": field, "message": message})

    select_cols = [*CREDENTIAL_IMPORT_COLUMNS, "row_number"]
    if "otp_seed" in df.columns:
        select_cols.append("otp_seed")
    rows = df.select(select_cols).to_dicts()

    dept_codes = {_to_str(r.get("dept_code")) for r in rows if _to_str(r.get("dept_code"))}
    dept_map: dict[str, UUID] = {}
    if dept_codes:
        dept_result = await db.execute(select(Department.code, Department.id).where(Department.code.in_(dept_codes)))
        dept_map = {str(code): dept_id for code, dept_id in dept_result.all()}

    normalized_rows: list[dict[str, Any]] = []
    pair_seen: set[tuple[str, str]] = set()
    pairs: set[tuple[UUID, str]] = set()

    for r in rows:
        row_number = int(r["row_number"])
        dept_code = _to_str(r.get("dept_code"))
        device_group = _normalize_enum(_to_str(r.get("device_group")))
        auth_type = _normalize_enum(_to_str(r.get("auth_type")) or AuthType.OTP_SEED.value)
        username = _to_str(r.get("username"))
        description = _to_str(r.get("description")) or None
        otp_seed = _to_str(r.get("otp_seed")) or None

        if not dept_code:
            add_error(row_number, "dept_code", "部门编码不能为空")
        if not device_group:
            add_error(row_number, "device_group", "设备分组不能为空")
        if not username:
            add_error(row_number, "username", "用户名不能为空")

        if device_group and device_group not in {e.value for e in DeviceGroup}:
            add_error(row_number, "device_group", f"设备分组无效: {device_group}")

        if auth_type not in {AuthType.OTP_SEED.value, AuthType.OTP_MANUAL.value}:
            add_error(row_number, "auth_type", f"认证类型无效: {auth_type}")

        if dept_code and device_group:
            key = (dept_code, device_group)
            if key in pair_seen:
                add_error(row_number, "device_group", "导入文件中 (dept_code, device_group) 重复")
            pair_seen.add(key)

        dept_id: UUID | None = None
        if dept_code:
            dept_id = dept_map.get(dept_code)
            if dept_id is None:
                add_error(row_number, "dept_code", f"部门不存在: {dept_code}")

        if dept_id is not None and device_group:
            pairs.add((dept_id, device_group))

        normalized_rows.append(
            {
                "row_number": row_number,
                "dept_code": dept_code or None,
                "dept_id": str(dept_id) if dept_id else None,
                "device_group": device_group or None,
                "auth_type": auth_type,
                "username": username or None,
                "otp_seed": otp_seed,
                "description": description,
            }
        )

    if pairs:
        dept_ids = {dept_id for dept_id, _ in pairs}
        groups = {g for _, g in pairs}
        existing_result = await db.execute(
            select(DeviceGroupCredential.dept_id, DeviceGroupCredential.device_group).where(
                and_(
                    DeviceGroupCredential.is_deleted.is_(False),
                    DeviceGroupCredential.dept_id.in_(dept_ids),
                    DeviceGroupCredential.device_group.in_(groups),
                )
            )
        )
        existing_pairs = {(dept_id, group) for dept_id, group in existing_result.all()}
        if (not allow_overwrite) and existing_pairs:
            for r in normalized_rows:
                if not r.get("dept_id") or not r.get("device_group"):
                    continue
                pair = (UUID(str(r["dept_id"])), str(r["device_group"]))
                if pair in existing_pairs:
                    add_error(int(r["row_number"]), "device_group", "分组凭据已存在")

    error_row_numbers = {int(e["row_number"]) for e in errors if int(e["row_number"]) > 0}
    valid_rows = [r for r in normalized_rows if r["row_number"] not in error_row_numbers]
    valid_df = pl.DataFrame(valid_rows)
    return valid_df, errors


async def persist_credentials(
    db: AsyncSession,
    valid_df: pl.DataFrame,
    *,
    allow_overwrite: bool = False,
) -> int:
    if valid_df.is_empty():
        return 0

    rows = valid_df.to_dicts()
    now = datetime.now().astimezone().replace(tzinfo=None)

    pairs: set[tuple[UUID, str]] = set()
    for r in rows:
        if r.get("dept_id") and r.get("device_group"):
            pairs.add((UUID(str(r["dept_id"])), str(r["device_group"])))

    existing_by_pair: dict[tuple[UUID, str], DeviceGroupCredential] = {}
    if pairs:
        dept_ids = {dept_id for dept_id, _ in pairs}
        groups = {g for _, g in pairs}
        existing_result = await db.execute(
            select(DeviceGroupCredential).where(
                and_(
                    DeviceGroupCredential.is_deleted.is_(False),
                    DeviceGroupCredential.dept_id.in_(dept_ids),
                    DeviceGroupCredential.device_group.in_(groups),
                )
            )
        )
        existing_by_pair = {(c.dept_id, c.device_group): c for c in existing_result.scalars().all()}

    to_insert: list[dict[str, Any]] = []
    to_update: list[tuple[DeviceGroupCredential, dict[str, Any]]] = []

    for r in rows:
        if not r.get("dept_id") or not r.get("device_group"):
            continue
        dept_id = UUID(str(r["dept_id"]))
        group = str(r["device_group"])
        otp_seed_encrypted: str | None = None
        if r.get("otp_seed"):
            otp_seed_encrypted = encrypt_credential(str(r["otp_seed"]), settings.NCM_CREDENTIAL_KEY)

        values: dict[str, Any] = {
            "dept_id": dept_id,
            "device_group": group,
            "auth_type": str(r.get("auth_type") or AuthType.OTP_SEED.value),
            "username": str(r.get("username") or ""),
            "otp_seed_encrypted": otp_seed_encrypted,
            "description": r.get("description"),
            "updated_at": now,
        }

        existing = existing_by_pair.get((dept_id, group))
        if existing is not None:
            if allow_overwrite:
                to_update.append((existing, values))
            continue

        values["id"] = uuid6.uuid7()
        values["version_id"] = uuid.uuid4().hex
        values["created_at"] = now
        to_insert.append(values)

    await db.flush()
    for existing, values in to_update:
        for k, v in values.items():
            setattr(existing, k, v)
    if to_insert:
        await db.execute(insert(DeviceGroupCredential), to_insert)
    await db.commit()

    return len(to_insert) + len(to_update)


async def export_credentials_df(db: AsyncSession) -> pl.DataFrame:
    result = await db.execute(
        select(DeviceGroupCredential, Department.code, Department.name)
        .join(Department, Department.id == DeviceGroupCredential.dept_id)
        .where(DeviceGroupCredential.is_deleted.is_(False))
        .order_by(DeviceGroupCredential.updated_at.desc())
    )
    rows: list[dict[str, Any]] = []
    for cred, dept_code, dept_name in result.all():
        rows.append(
            {
                "dept_code": str(dept_code),
                "dept_name": str(dept_name),
                "device_group": cred.device_group,
                "auth_type": cred.auth_type,
                "username": cred.username,
                "has_otp_seed": bool(cred.otp_seed_encrypted),
                "description": cred.description or "",
                "created_at": cred.created_at.isoformat() if cred.created_at else "",
                "updated_at": cred.updated_at.isoformat() if cred.updated_at else "",
            }
        )
    return pl.DataFrame(rows)


def build_credential_import_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook.active is None")
    ws = cast(Worksheet, ws)
    ws.title = "credentials"

    header_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(CREDENTIAL_IMPORT_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(14, min(30, len(col_name) + 6))

    samples = [
        {
            "dept_code": "HQ",
            "device_group": "core",
            "auth_type": "otp_seed",
            "username": "admin",
            "otp_seed": "JBSWY3DPEHPK3PXP",
            "description": "示例：OTP 种子凭据",
        },
        {
            "dept_code": "HQ",
            "device_group": "access",
            "auth_type": "otp_manual",
            "username": "admin",
            "otp_seed": "",
            "description": "示例：OTP 手动输入（无需种子）",
        },
    ]
    for row_idx, sample in enumerate(samples, start=2):
        for col_idx, col_name in enumerate(CREDENTIAL_IMPORT_TEMPLATE_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=sample.get(col_name, ""))

    dv_group = DataValidation(type="list", formula1=f'"{",".join([e.value for e in DeviceGroup])}"', allow_blank=False)
    dv_auth = DataValidation(
        type="list", formula1=f'"{AuthType.OTP_SEED.value},{AuthType.OTP_MANUAL.value}"', allow_blank=False
    )

    group_col = CREDENTIAL_IMPORT_TEMPLATE_COLUMNS.index("device_group") + 1
    auth_col = CREDENTIAL_IMPORT_TEMPLATE_COLUMNS.index("auth_type") + 1

    for dv, col in ((dv_group, group_col), (dv_auth, auth_col)):
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=col).coordinate}:{ws.cell(row=500, column=col).coordinate}")

    ws.freeze_panes = "A2"
    wb.save(path)
