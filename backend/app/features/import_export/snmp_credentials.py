"""
@Author: li
@Email: lijianqiao2906@live.com
@FileName: snmp_credentials.py
@DateTime: 2026/01/16
@Docs: SNMP 凭据导入导出（部门维度）
"""

import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, cast
from uuid import UUID

import polars as pl
import uuid6
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import and_, insert, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.encryption import encrypt_credential
from app.core.enums import SnmpAuthProtocol, SnmpPrivProtocol, SnmpSecurityLevel, SnmpVersion
from app.models.dept import Department
from app.models.snmp_credential import DeptSnmpCredential

SNMP_CRED_IMPORT_COLUMNS: list[str] = [
    "dept_code",
    "snmp_version",
    "port",
    "community",
    "v3_username",
    "v3_security_level",
    "v3_auth_proto",
    "v3_auth_key",
    "v3_priv_proto",
    "v3_priv_key",
    "description",
]

SNMP_CRED_IMPORT_TEMPLATE_COLUMNS: list[str] = [
    *SNMP_CRED_IMPORT_COLUMNS,
]

SNMP_CRED_IMPORT_COLUMN_ALIASES: dict[str, str] = {
    "部门编码": "dept_code",
    "dept_code": "dept_code",
    "版本": "snmp_version",
    "snmp_version": "snmp_version",
    "端口": "port",
    "port": "port",
    "团体字": "community",
    "community": "community",
    "v3用户名": "v3_username",
    "v3_username": "v3_username",
    "安全级别": "v3_security_level",
    "v3_security_level": "v3_security_level",
    "认证协议": "v3_auth_proto",
    "v3_auth_proto": "v3_auth_proto",
    "认证密钥": "v3_auth_key",
    "v3_auth_key": "v3_auth_key",
    "加密协议": "v3_priv_proto",
    "v3_priv_proto": "v3_priv_proto",
    "加密密钥": "v3_priv_key",
    "v3_priv_key": "v3_priv_key",
    "描述": "description",
    "说明": "description",
    "description": "description",
}


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize_enum(value: str) -> str:
    return value.strip().lower().replace(" ", "_")


def _parse_int(value: str, default: int | None = None) -> int | None:
    v = value.strip()
    if not v:
        return default
    try:
        return int(float(v))
    except Exception:
        return None


async def validate_snmp_credentials(
    db: AsyncSession,
    df: pl.DataFrame,
    *,
    allow_overwrite: bool = False,
) -> tuple[pl.DataFrame, list[dict[str, Any]]]:
    errors: list[dict[str, Any]] = []

    for col in SNMP_CRED_IMPORT_COLUMNS:
        if col not in df.columns:
            errors.append({"row_number": 0, "field": col, "message": f"缺少列: {col}"})

    if errors:
        return pl.DataFrame(), errors

    def add_error(row_number: int, field: str | None, message: str) -> None:
        errors.append({"row_number": int(row_number), "field": field, "message": message})

    rows = df.select([*SNMP_CRED_IMPORT_COLUMNS, "row_number"]).to_dicts()

    dept_codes = {_to_str(r.get("dept_code")) for r in rows if _to_str(r.get("dept_code"))}
    dept_map: dict[str, UUID] = {}
    if dept_codes:
        dept_result = await db.execute(select(Department.code, Department.id).where(Department.code.in_(dept_codes)))
        dept_map = {str(code): dept_id for code, dept_id in dept_result.all()}

    normalized_rows: list[dict[str, Any]] = []
    dept_seen: set[str] = set()
    dept_ids: set[UUID] = set()

    for r in rows:
        row_number = int(r["row_number"])
        dept_code = _to_str(r.get("dept_code"))
        snmp_version = _normalize_enum(_to_str(r.get("snmp_version")) or SnmpVersion.V2C.value)
        port = _parse_int(_to_str(r.get("port")), default=161)
        community = _to_str(r.get("community")) or None
        v3_username = _to_str(r.get("v3_username")) or None
        v3_security_level = _normalize_enum(
            _to_str(r.get("v3_security_level")) or SnmpSecurityLevel.NO_AUTH_NO_PRIV.value
        )
        v3_auth_proto = _normalize_enum(_to_str(r.get("v3_auth_proto")) or SnmpAuthProtocol.SHA.value)
        v3_auth_key = _to_str(r.get("v3_auth_key")) or None
        v3_priv_proto = _normalize_enum(_to_str(r.get("v3_priv_proto")) or SnmpPrivProtocol.AES.value)
        v3_priv_key = _to_str(r.get("v3_priv_key")) or None
        description = _to_str(r.get("description")) or None

        if not dept_code:
            add_error(row_number, "dept_code", "部门编码不能为空")

        if dept_code:
            if dept_code in dept_seen:
                add_error(row_number, "dept_code", "导入文件中 dept_code 重复")
            dept_seen.add(dept_code)

        if snmp_version not in {e.value for e in SnmpVersion}:
            add_error(row_number, "snmp_version", f"SNMP 版本无效: {snmp_version}")

        if port is None or not (1 <= int(port) <= 65535):
            add_error(row_number, "port", "端口无效")

        if snmp_version == SnmpVersion.V2C.value:
            if not community:
                add_error(row_number, "community", "SNMP v2c 必须填写 community")

        if snmp_version == SnmpVersion.V3.value:
            if not v3_username:
                add_error(row_number, "v3_username", "SNMP v3 必须填写 v3_username")
            if v3_security_level not in {e.value for e in SnmpSecurityLevel}:
                add_error(row_number, "v3_security_level", f"安全级别无效: {v3_security_level}")
            if v3_auth_proto and v3_auth_proto not in {e.value for e in SnmpAuthProtocol}:
                add_error(row_number, "v3_auth_proto", f"认证协议无效: {v3_auth_proto}")
            if v3_priv_proto and v3_priv_proto not in {e.value for e in SnmpPrivProtocol}:
                add_error(row_number, "v3_priv_proto", f"加密协议无效: {v3_priv_proto}")

            if v3_security_level in {SnmpSecurityLevel.AUTH_NO_PRIV.value, SnmpSecurityLevel.AUTH_PRIV.value}:
                if not v3_auth_key:
                    add_error(row_number, "v3_auth_key", "安全级别要求填写 v3_auth_key")
            if v3_security_level == SnmpSecurityLevel.AUTH_PRIV.value:
                if not v3_priv_key:
                    add_error(row_number, "v3_priv_key", "安全级别要求填写 v3_priv_key")

        dept_id: UUID | None = None
        if dept_code:
            dept_id = dept_map.get(dept_code)
            if dept_id is None:
                add_error(row_number, "dept_code", f"部门不存在: {dept_code}")
            else:
                dept_ids.add(dept_id)

        normalized_rows.append(
            {
                "row_number": row_number,
                "dept_code": dept_code or None,
                "dept_id": str(dept_id) if dept_id else None,
                "snmp_version": snmp_version,
                "port": int(port) if port is not None else None,
                "community": community,
                "v3_username": v3_username,
                "v3_security_level": v3_security_level,
                "v3_auth_proto": v3_auth_proto,
                "v3_auth_key": v3_auth_key,
                "v3_priv_proto": v3_priv_proto,
                "v3_priv_key": v3_priv_key,
                "description": description,
            }
        )

    if dept_ids:
        existing_result = await db.execute(
            select(DeptSnmpCredential.dept_id).where(
                and_(DeptSnmpCredential.is_deleted.is_(False), DeptSnmpCredential.dept_id.in_(dept_ids))
            )
        )
        existing_ids = set(existing_result.scalars().all())
        if (not allow_overwrite) and existing_ids:
            for r in normalized_rows:
                if not r.get("dept_id"):
                    continue
                if UUID(str(r["dept_id"])) in existing_ids:
                    add_error(int(r["row_number"]), "dept_code", "SNMP 凭据已存在")

    error_row_numbers = {int(e["row_number"]) for e in errors if int(e["row_number"]) > 0}
    valid_rows = [r for r in normalized_rows if r["row_number"] not in error_row_numbers]
    valid_df = pl.DataFrame(valid_rows)
    return valid_df, errors


async def persist_snmp_credentials(
    db: AsyncSession,
    valid_df: pl.DataFrame,
    *,
    allow_overwrite: bool = False,
) -> int:
    if valid_df.is_empty():
        return 0

    rows = valid_df.to_dicts()
    now = datetime.now().astimezone().replace(tzinfo=None)

    dept_ids: set[UUID] = set()
    for r in rows:
        if r.get("dept_id"):
            dept_ids.add(UUID(str(r["dept_id"])))

    existing_by_dept: dict[UUID, DeptSnmpCredential] = {}
    if dept_ids:
        existing_result = await db.execute(
            select(DeptSnmpCredential).where(
                and_(DeptSnmpCredential.is_deleted.is_(False), DeptSnmpCredential.dept_id.in_(dept_ids))
            )
        )
        existing_by_dept = {c.dept_id: c for c in existing_result.scalars().all()}

    to_insert: list[dict[str, Any]] = []
    to_update: list[tuple[DeptSnmpCredential, dict[str, Any]]] = []

    for r in rows:
        if not r.get("dept_id"):
            continue
        dept_id = UUID(str(r["dept_id"]))
        snmp_version = str(r.get("snmp_version") or SnmpVersion.V2C.value)

        community_encrypted: str | None = None
        if r.get("community"):
            community_encrypted = encrypt_credential(str(r["community"]), settings.NCM_CREDENTIAL_KEY)

        v3_auth_key_encrypted: str | None = None
        if r.get("v3_auth_key"):
            v3_auth_key_encrypted = encrypt_credential(str(r["v3_auth_key"]), settings.NCM_CREDENTIAL_KEY)

        v3_priv_key_encrypted: str | None = None
        if r.get("v3_priv_key"):
            v3_priv_key_encrypted = encrypt_credential(str(r["v3_priv_key"]), settings.NCM_CREDENTIAL_KEY)

        values: dict[str, Any] = {
            "dept_id": dept_id,
            "snmp_version": snmp_version,
            "port": int(r.get("port") or 161),
            "community_encrypted": community_encrypted,
            "v3_username": r.get("v3_username"),
            "v3_security_level": r.get("v3_security_level"),
            "v3_auth_proto": r.get("v3_auth_proto"),
            "v3_auth_key_encrypted": v3_auth_key_encrypted,
            "v3_priv_proto": r.get("v3_priv_proto"),
            "v3_priv_key_encrypted": v3_priv_key_encrypted,
            "description": r.get("description"),
            "updated_at": now,
        }

        existing = existing_by_dept.get(dept_id)
        if existing is not None:
            if allow_overwrite:
                to_update.append((existing, values))
            continue

        values["id"] = uuid6.uuid7()
        values["version_id"] = uuid.uuid4().hex
        values["created_at"] = now
        to_insert.append(values)

    await db.flush()
    for existing, values in to_update:
        for k, v in values.items():
            setattr(existing, k, v)
    if to_insert:
        await db.execute(insert(DeptSnmpCredential), to_insert)
    await db.commit()

    return len(to_insert) + len(to_update)


async def export_snmp_credentials_df(db: AsyncSession) -> pl.DataFrame:
    result = await db.execute(
        select(DeptSnmpCredential, Department.code, Department.name)
        .join(Department, Department.id == DeptSnmpCredential.dept_id)
        .where(DeptSnmpCredential.is_deleted.is_(False))
        .order_by(DeptSnmpCredential.updated_at.desc())
    )
    rows: list[dict[str, Any]] = []
    for cred, dept_code, dept_name in result.all():
        rows.append(
            {
                "dept_code": str(dept_code),
                "dept_name": str(dept_name),
                "snmp_version": cred.snmp_version,
                "port": cred.port,
                "has_community": bool(cred.community_encrypted),
                "v3_username": cred.v3_username or "",
                "v3_security_level": cred.v3_security_level or "",
                "v3_auth_proto": cred.v3_auth_proto or "",
                "has_v3_auth_key": bool(cred.v3_auth_key_encrypted),
                "v3_priv_proto": cred.v3_priv_proto or "",
                "has_v3_priv_key": bool(cred.v3_priv_key_encrypted),
                "description": cred.description or "",
                "created_at": cred.created_at.isoformat() if cred.created_at else "",
                "updated_at": cred.updated_at.isoformat() if cred.updated_at else "",
            }
        )
    return pl.DataFrame(rows)


def build_snmp_credential_import_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook.active is None")
    ws = cast(Worksheet, ws)
    ws.title = "snmp_credentials"

    header_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(SNMP_CRED_IMPORT_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(14, min(30, len(col_name) + 6))

    samples = [
        {
            "dept_code": "HQ",
            "snmp_version": "v2c",
            "port": "161",
            "community": "public",
            "v3_username": "",
            "v3_security_level": "",
            "v3_auth_proto": "",
            "v3_auth_key": "",
            "v3_priv_proto": "",
            "v3_priv_key": "",
            "description": "示例：SNMP v2c",
        },
        {
            "dept_code": "HQ",
            "snmp_version": "v3",
            "port": "161",
            "community": "",
            "v3_username": "snmpuser",
            "v3_security_level": "auth_priv",
            "v3_auth_proto": "sha",
            "v3_auth_key": "auth-key",
            "v3_priv_proto": "aes",
            "v3_priv_key": "priv-key",
            "description": "示例：SNMP v3",
        },
    ]

    for row_idx, sample in enumerate(samples, start=2):
        for col_idx, col_name in enumerate(SNMP_CRED_IMPORT_TEMPLATE_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=sample.get(col_name, ""))

    dv_version = DataValidation(
        type="list", formula1=f'"{",".join([e.value for e in SnmpVersion])}"', allow_blank=False
    )
    dv_sec = DataValidation(
        type="list", formula1=f'"{",".join([e.value for e in SnmpSecurityLevel])}"', allow_blank=True
    )
    dv_auth = DataValidation(
        type="list", formula1=f'"{",".join([e.value for e in SnmpAuthProtocol])}"', allow_blank=True
    )
    dv_priv = DataValidation(
        type="list", formula1=f'"{",".join([e.value for e in SnmpPrivProtocol])}"', allow_blank=True
    )

    for dv, col_name in (
        (dv_version, "snmp_version"),
        (dv_sec, "v3_security_level"),
        (dv_auth, "v3_auth_proto"),
        (dv_priv, "v3_priv_proto"),
    ):
        col = SNMP_CRED_IMPORT_TEMPLATE_COLUMNS.index(col_name) + 1
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=col).coordinate}:{ws.cell(row=500, column=col).coordinate}")

    ws.freeze_panes = "A2"
    wb.save(path)
