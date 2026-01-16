"""Parse tabular files (CSV/Excel) into Polars DataFrame.

将 CSV/Excel 文件解析为 Polars DataFrame。

This module focuses on converting uploaded files into a normalized tabular
representation for further validation and persistence.

该模块侧重于将上传的文件转换为规范化的表格用于进一步验证和持久化的表示。

Key behaviors / 关键行为:
    - Supports CSV and Excel (.xlsx/.xlsm/.xls).
      支持 CSV 与 Excel。
    - Adds a `row_number` column starting from 1 (data rows only).
      自动添加 `row_number`（从 1 开始，表示数据行号，不含表头）。
    - Casts all columns to string (Utf8) with `strict=False` to avoid schema surprises.
      所有列统一转为字符串（Utf8，strict=False）以减少类型差异带来的问题。

Excel parsing strategy / Excel 解析策略:
    - Prefer `polars.read_excel` when available.
      优先使用 `polars.read_excel`（若可用）。
    - Fallback to `openpyxl` to avoid pandas→polars requiring pyarrow.
      回退到 `openpyxl`，避免 pandas→polars 需要 pyarrow 的问题。

Examples:
    >>> from pathlib import Path
    >>> from fastapi_import_export.parse import parse_tabular_file
    >>> parsed = parse_tabular_file(Path("devices.csv"), filename="devices.csv")
    >>> "row_number" in parsed.df.columns
    True
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

import polars as pl
from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class ParsedTable:
    """Parsed tabular file result.

    解析后的表格文件结果。

    Attributes:
        df: Parsed data as a Polars DataFrame (includes `row_number`).
            解析后的 Polars DataFrame（包含 `row_number`）。
        total_rows: Number of data rows.
            数据行数。
        columns: Column names list.
            列名列表。
    """

    df: pl.DataFrame
    total_rows: int
    columns: list[str]


def _read_excel_to_polars(file_path: Path, *, sheet_name: str | int | None = None) -> pl.DataFrame:
    """Read an Excel file into Polars DataFrame.

    将 Excel 文件读取为 Polars DataFrame。

    Args:
        file_path: Excel file path.
            Excel 文件路径。
        sheet_name: Sheet selector (None for first, int index, or sheet name).
            工作表选择器（None 表示第一个；int 表示索引；str 表示工作表名）。

    Returns:
        Polars DataFrame.
            返回 Polars DataFrame。

    Raises:
        KeyError: If a given sheet name does not exist.
            当 sheet_name 为不存在的工作表名时抛出。
        IndexError: If a given sheet index is out of range.
            当 sheet_name 为越界索引时抛出。
    """
    try:
        read_excel = getattr(pl, "read_excel", None)
        if callable(read_excel):
            df = read_excel(str(file_path), sheet_name=sheet_name)  # type: ignore[call-arg]
            return cast(pl.DataFrame, df)
    except Exception:
        pass

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            ws = wb.worksheets[0]
        elif isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return pl.DataFrame()

        headers: list[str] = []
        seen: dict[str, int] = {}
        for idx, h in enumerate(header_row, start=1):
            name = str(h).strip() if h is not None and str(h).strip() else f"column_{idx}"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            headers.append(name)

        records: list[dict[str, Any]] = []
        for row in rows_iter:
            if row is None:
                continue
            record = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            records.append(record)

        return pl.DataFrame(records)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def parse_tabular_file(file_path: Path, *, filename: str) -> ParsedTable:
    """Parse a CSV/Excel file to `ParsedTable`.

    将 CSV/Excel 文件解析为 `ParsedTable`。

    Args:
        file_path: Path to the file on disk.
            文件磁盘路径。
        filename: Original filename (used for suffix detection).
            原始文件名（用于判断扩展名）。

    Returns:
        ParsedTable: Parsed result with `df/total_rows/columns`.
            解析结果（包含 df/total_rows/columns）。

    Raises:
        ValueError: If the file type is not supported.
            不支持的文件类型时抛出。

    Examples:
        >>> from pathlib import Path
        >>> t = parse_tabular_file(Path("devices.xlsx"), filename="devices.xlsx")
        >>> t.total_rows >= 0
        True
    """
    suffix = Path(filename).suffix.lower().lstrip(".")
    if suffix in {"csv"}:
        df = pl.read_csv(str(file_path), infer_schema=False, encoding="utf8-lossy")
    elif suffix in {"xlsx", "xlsm", "xls"}:
        df = _read_excel_to_polars(file_path)
    else:
        raise ValueError(f"不支持的文件类型: .{suffix}")

    df = df.with_columns([pl.all().cast(pl.Utf8, strict=False)])
    df = df.with_row_index(name="row_number", offset=1)
    return ParsedTable(df=df, total_rows=df.height, columns=list(df.columns))


def normalize_columns(df: pl.DataFrame, column_mapping: dict[str, str]) -> pl.DataFrame:
    """Normalize column names using a mapping table.

    基于列名映射表标准化 DataFrame 列名。

    Args:
        df: Input DataFrame.
            输入 DataFrame。
        column_mapping: Mapping from raw header to canonical header.
            列名映射（原始表头 -> 规范表头）。

    Returns:
        Renamed DataFrame.
            返回重命名后的 DataFrame。
    """
    normalized: dict[str, str] = {}
    for c in df.columns:
        c_norm = str(c).strip()
        normalized[c] = column_mapping.get(c_norm, c_norm)
    return df.rename(normalized)


def dataframe_to_preview_rows(df: pl.DataFrame) -> list[dict[str, Any]]:
    """Convert a DataFrame to preview rows (list of dict).

    将 DataFrame 转换为预览行（列表 of dict）。

    Args:
        df: Polars DataFrame.
            Polars DataFrame。

    Returns:
        List of rows as dicts (column -> value).
            行列表（dict：列名 -> 值）。
    """
    return df.to_dicts()
